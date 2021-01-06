import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'multiply'
wb.save('multiplicationTables')

wb = openpyxl.load_workbook('multiplicationTable.xlsx')
sheet = wb['multiply']
number = int(input("Enter a number to create a multiplication table - "))

column_letter = '' #column letter is to extract the letter of the column
counter = 1 #counter is to help print the the first row labels
i = 1 #a counter to help print the products for the table

# to print the first column of labels
for r in range(2,number + 2):
    sheet['A' + str(r)].font = Font(bold = True)
    sheet['A'+ str(r)] = r - 1

# to print the rest of the table
for column in range(2,number + 2):
    column_letter = get_column_letter(column)
    i = 1
    for row in range(1,number + 2):
        if(row == 1):
            sheet[column_letter + str(row)].font = Font(bold = True)
            sheet[column_letter + str(row)] = counter
            counter += 1
            
        else:
            sheet[column_letter + str(row)] = i * (column - 1)
            i += 1
        
wb.save('multiplicationTable.xlsx')
